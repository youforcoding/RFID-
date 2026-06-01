import serial
import pandas as pd
from datetime import datetime
import sys
import time
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

PORT      = "COM3"
BAUD_RATE = 9600
FILE      = "presence.xlsx"
DB_FILE   = "presence.db"

students = {
    "9C9E1A06": ("étudiant",  "1"),
    "C0861217": ("étudiant",    "2"),
    "B0E9C917": ("étudiant",  "3"),
    "C01DED17": ("étudiant",   "4"),
    "C0067217": ("étudiant",   "5"),
}

# = SQLITE INIT 
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS presence (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            uid      TEXT,
            nom      TEXT,
            prenom   TEXT,
            heure    TEXT,
            statut   TEXT,
            date     TEXT
        )
    """)
    # Table pour professor acc
    c.execute("""
        CREATE TABLE IF NOT EXISTS professeurs (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT
        )
    """)
    # compte de prof  (username: prof  /  password: ****** )
    c.execute("INSERT OR IGNORE INTO professeurs (username, password) VALUES (?, ?)",
              ("prof", "******"))
    conn.commit()
    conn.close()

def init_today():
    """Insert absent rows for all students if not already done today."""
    today = datetime.now().strftime("%Y-%m-%d")
    conn  = sqlite3.connect(DB_FILE)
    c     = conn.cursor()
    for uid, (nom, prenom) in students.items():
        c.execute("SELECT id FROM presence WHERE uid=? AND date=?", (uid, today))
        if not c.fetchone():
            c.execute(
                "INSERT INTO presence (uid, nom, prenom, heure, statut, date) VALUES (?,?,?,?,?,?)",
                (uid, nom, prenom, "-", "Absent", today)
            )
    conn.commit()
    conn.close()

def mark_present(uid):
    today = datetime.now().strftime("%Y-%m-%d")
    heure = datetime.now().strftime("%H:%M:%S")
    conn  = sqlite3.connect(DB_FILE)
    c     = conn.cursor()
    c.execute(
        "UPDATE presence SET heure=?, statut='Présent' WHERE uid=? AND date=?",
        (heure, uid, today)
    )
    conn.commit()
    conn.close()
    return heure

def get_today_data():
    today = datetime.now().strftime("%Y-%m-%d")
    conn  = sqlite3.connect(DB_FILE)
    c     = conn.cursor()
    c.execute(
        "SELECT uid, nom, prenom, heure, statut FROM presence WHERE date=?",
        (today,)
    )
    rows = c.fetchall()
    conn.close()
    return rows

#  BORDER HELPER 
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

#  SAVE EXCEL 
def save_excel():
    rows = get_today_data()
    df   = pd.DataFrame(rows, columns=["UID", "Nom", "Prénom", "Heure", "Statut"])
    df.to_excel(FILE, index=False)

    wb = load_workbook(FILE)
    ws = wb.active

    # TITRE
    ws.insert_rows(1)
    titre = "LISTE DE PRÉSENCE  —  " + datetime.now().strftime("%d/%m/%Y")
    ws["A1"] = titre
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 36
    ws["A1"].font      = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A3A5C")

    # SOUS-TITRE stats
    present_count = sum(1 for r in rows if r[4] == "Présent")
    absent_count  = len(rows) - present_count
    ws.insert_rows(2)
    ws["A2"] = f"  ✅ Présents : {present_count}     ❌ Absents : {absent_count}     👥 Total : {len(rows)}"
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 22
    ws["A2"].font      = Font(name="Calibri", size=11, italic=True, color="1A3A5C")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill      = PatternFill("solid", fgColor="D6E4F0")

    # HEADER
    ws.row_dimensions[3].height = 22
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws[f"{col}3"]
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

    # COLUMN WIDTH
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # ROW STYLES
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        statut = row[4].value
        if statut == "Présent":
            bg, statut_bg = "C6EFCE", "70AD47"
        else:
            bg, statut_bg = "FCE4EC", "FF4D6D"

        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(name="Calibri", size=11)
            cell.border    = thin_border()
            cell.fill      = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row[0].row].height = 20
        statut_cell      = row[4]
        statut_cell.fill = PatternFill("solid", fgColor=statut_bg)
        statut_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    wb.save(FILE)
    print(f"[EXCEL] Sauvegardé — Présents: {present_count}/{len(rows)}")

#= MAIN 
init_db()
init_today()
save_excel()

ser = serial.Serial(PORT, BAUD_RATE, timeout=2)
time.sleep(2)
print("[OK] COM CONNECTÉ")
print("SYSTÈME RFID PRÊT")

present_today = set()
ser.reset_input_buffer()

while True:
    try:
        uid = ser.readline().decode(errors='ignore').strip().upper()
        uid = uid.replace("UID:", "").strip()
        if not uid:
            continue

        print("[CARTE]", uid)

        if uid in students:
            nom, prenom = students[uid]
            if uid not in present_today:
                present_today.add(uid)
                mark_present(uid)
                save_excel()
                ser.write(("OK " + nom + " " + prenom + "\n").encode())
            else:
                ser.write(b"DEJA\n")
        else:
            ser.write(b"UNKNOWN\n")

    except Exception as e:
        print("[ERREUR]", e)
        time.sleep(1)
