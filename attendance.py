import serial
import pandas as pd
from datetime import datetime
import sys
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

PORT = "COM3"
BAUD_RATE = 9600
FILE = "presence.xlsx"

students = {
    "9C9E1A06": ("Younes",  "Aoudache"),
    "C0861217": ("Zaki",    "Aissat"),
    "B0E9C917": ("Yacine",  "Berkache"),
    "C01DED17": ("Samia",   "Benslimane"),
    "C0067217": ("Rania",   "Khemissat"),
}

ser = serial.Serial(PORT, BAUD_RATE, timeout=2)
time.sleep(2)

print("[OK] COM CONNECTÉ")
print("SYSTÈME RFID PRÊT")

present_today = set()

data = [[uid, nom, prenom, "-", "Absent"] for uid, (nom, prenom) in students.items()]

# ================= BORDER HELPER =================
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ================= SAVE EXCEL =================
def save_excel():
    df = pd.DataFrame(data, columns=["UID", "Nom", "Prénom", "Heure", "Statut"])
    df.to_excel(FILE, index=False)

    wb = load_workbook(FILE)
    ws = wb.active

    # ===== TITRE =====
    ws.insert_rows(1)
    titre = "LISTE DE PRÉSENCE  —  " + datetime.now().strftime("%d/%m/%Y")
    ws["A1"] = titre
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 36

    ws["A1"].font      = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A3A5C")

    # ===== SOUS-TITRE stats =====
    present_count = sum(1 for row in data if row[4] == "Présent")
    absent_count  = len(data) - present_count
    ws.insert_rows(2)
    ws["A2"] = f"  ✅ Présents : {present_count}     ❌ Absents : {absent_count}     👥 Total : {len(data)}"
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 22
    ws["A2"].font      = Font(name="Calibri", size=11, italic=True, color="1A3A5C")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill      = PatternFill("solid", fgColor="D6E4F0")

    # ===== HEADER =====
    ws.row_dimensions[3].height = 22
    header_cols = ["A", "B", "C", "D", "E"]
    for col in header_cols:
        cell = ws[f"{col}3"]
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

    # ===== COLUMN WIDTH =====
    ws.column_dimensions["A"].width = 16   # UID
    ws.column_dimensions["B"].width = 18   # Nom
    ws.column_dimensions["C"].width = 18   # Prénom
    ws.column_dimensions["D"].width = 14   # Heure
    ws.column_dimensions["E"].width = 14   # Statut

    # ===== ROW STYLES =====
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=0):
        statut = row[4].value

        # alternating background base
        if statut == "Présent":
            bg = "C6EFCE"   # vert clair
            fg = "276221"   # vert foncé (texte statut)
            statut_bg = "70AD47"
        else:
            bg = "FCE4EC"   # rouge clair
            fg = "C62828"
            statut_bg = "FF4D6D"

        for j, cell in enumerate(row):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(name="Calibri", size=11)
            cell.border    = thin_border()
            cell.fill      = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row[0].row].height = 20

        # Statut cell special style
        statut_cell = row[4]
        statut_cell.fill = PatternFill("solid", fgColor=statut_bg)
        statut_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    wb.save(FILE)
    print(f"[EXCEL] Sauvegardé — Présents: {present_count}/{len(data)}")

# ================= INIT =================
save_excel()
ser.reset_input_buffer()

# ================= LOOP =================
while True:
    try:
        uid = ser.readline().decode(errors='ignore').strip().upper()
        uid = uid.replace("UID:", "").strip()

        if not uid:
            continue

        print("[CARTE]", uid)

        if uid in students:
            nom, prenom = students[uid]

            if uid not in present_today:
                present_today.add(uid)
                heure = datetime.now().strftime("%H:%M:%S")

                for i in range(len(data)):
                    if data[i][0] == uid:
                        data[i] = [uid, nom, prenom, heure, "Présent"]

                save_excel()
                ser.write(("OK " + nom + " " + prenom + "\n").encode())

            else:
                ser.write(b"DEJA\n")

        else:
            ser.write(b"UNKNOWN\n")

    except Exception as e:
        print("[ERREUR]", e)
        time.sleep(1)
